from __future__ import annotations

import os
import sys
import json
import glob
import sqlite3
import logging
import logging.handlers
import re
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple

import pandas as pd
from pydantic import BaseModel, Field, ValidationError
from pypdf import PdfReader
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Google GenAI
# ---------------------------------------------------------------------------
try:
    from google import genai
    from google.genai import types
    GENAI_DISPONIBLE = True
except ImportError:
    GENAI_DISPONIBLE = False

# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


# ===========================================================================
# 1. CONFIGURACION CENTRALIZADA
# ===========================================================================
@dataclass
class Config:
    gemini_api_key: str = field(default_factory=lambda: os.getenv("GEMINI_API_KEY", ""))
    modelo_llm: str = "gemini-3.6-flash"
    directorio_entrada: str = "facturas"
    directorio_salida: str = "salida"
    db_path: str = "sistema_facturas.db"
    json_historial: str = "historial_facturas_comex.json"
    log_path: str = "logs/facturador.log"
    log_nivel_consola: str = "INFO"
    log_nivel_archivo: str = "DEBUG"
    tolerancia_redondeo: float = 0.05
    tolerancia_advertencia: float = 0.01
    timeout_llm: int = 120  # segundos

    def __post_init__(self):
        Path(self.directorio_salida).mkdir(parents=True, exist_ok=True)
        Path("logs").mkdir(parents=True, exist_ok=True)
        if not self.gemini_api_key or self.gemini_api_key == "tu_clave_de_gemini_aqui":
            raise ValueError("Configura GEMINI_API_KEY en el archivo .env")


# ===========================================================================
# 2. LOGGING PROFESIONAL
# ===========================================================================
def setup_logging(cfg: Config) -> logging.Logger:
    logger = logging.getLogger("facturador")
    logger.setLevel(logging.DEBUG)
    logger.handlers = []

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(getattr(logging, cfg.log_nivel_consola.upper()))
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    fh = logging.handlers.RotatingFileHandler(
        cfg.log_path, maxBytes=5_000_000, backupCount=3, encoding="utf-8"
    )
    fh.setLevel(getattr(logging, cfg.log_nivel_archivo.upper()))
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


# ===========================================================================
# 3. MODELOS DE DATOS (Pydantic)
# ===========================================================================
class EstadoValidacion(str, Enum):
    CORRECTO = "CORRECTO"
    ADVERTENCIA = "ADVERTENCIA"
    ERROR = "ERROR"


class ResultadoValidacionItem(BaseModel):
    estado: EstadoValidacion = EstadoValidacion.CORRECTO
    total_calculado: float = 0.0
    diferencia: float = 0.0
    mensaje: str = ""


class ItemFactura(BaseModel):
    codigo_producto: str = Field(description="Codigo del producto. Ej: F243858000, F130802000.")
    descripcion_original: str = Field(default="", description="Descripcion tal como viene en el PDF.")
    descripcion_tecnica_limpia: str = Field(
        default="",
        description=(
            "Descripcion tecnica limpia SIN metadatos. "
            "Ej: 'WINDOW,TEFLON (1673MA)', 'O-RING 234 FKM 95 EL07 BLUE-DOT', "
            "'BEARING DRIVESHAFT UPR 1673MA', 'INSTMT 1677EA ACOUSTIC 8 CHANNEL ELECNS'."
        )
    )
    cantidad: float = Field(description="Cantidad numerica.")
    unidad_medida: str = Field(default="EA", description="Unidad: EA, BAG, KG, LB, PCS, SET, etc.")
    precio_unitario: float = Field(description="Precio unitario.")
    valor_total_item: float = Field(description="Extended Price / Total de la linea.")
    porcentaje: Optional[float] = Field(default=None, description="Porcentaje cuando aplique.")
    numero_referencia: str = Field(
        default="NO TIENE DELIVERY",
        description="Numero de Delivery si existe. Ej: 824273721, 824307893."
    )
    numero_orden: str = Field(
        default="NO TIENE ORDEN",
        description="Numero de Orden/PO si existe. Ej: 4200362929, 4513545869."
    )
    numero_serie: str = Field(
        default="NO TIENE SERIAL",
        description="Serial No. si existe. Ej: 10264520, 10332647."
    )
    codigo_exportacion: str = Field(default="NO TIENE CODIGO DE EXPORTACION")
    codigo_importacion: str = Field(default="NO TIENE CODIGO DE IMPORTACION")
    pais_origen: str = Field(default="N/A", description="Country of Origin. Ej: USA, Mexico, Thailand.")
    customs_import_mexico: str = Field(
        default="NO APLICA",
        description="MEXICAN CUSTOMS IMPORT number si existe. Ej: 214737891001336."
    )
    validacion: ResultadoValidacionItem = Field(default_factory=ResultadoValidacionItem)


class EncabezadoFactura(BaseModel):
    numero_factura: str = Field(description="Commercial Invoice Number. Ej: 4634461, 4638446.")
    fecha_factura: str = Field(default="N/A", description="Fecha de emision.")
    moneda: str = Field(default="USD", description="Moneda.")
    emisor_usa: str = Field(description="Emisor. Ej: Baker Hughes Oilfield Operations LLC.")
    comprador_colombia: str = Field(description="Comprador en Colombia.")
    modo_transporte: str = Field(default="N/A", description="Mode of Transport. Ej: Air, Sea.")
    agente_forwarder: str = Field(default="N/A", description="Forwarding Agent. Ej: DHL EXPRESS.")
    numero_orden: str = Field(default="NO DISPONIBLE", description="Numero de orden/PO general si existe.")
    pais_destino_final: str = Field(default="Colombia", description="Country of Ultimate Destination.")


class PieFactura(BaseModel):
    peso_bruto_total_lb: str = Field(default="N/A")
    peso_bruto_total_kg: str = Field(default="N/A")
    peso_neto_total_lb: str = Field(default="N/A")
    peso_neto_total_kg: str = Field(default="N/A")
    valor_neto: float = Field(default=0.0, description="Net Value = suma pura de Extended Price.")
    freight_inland: float = Field(default=0.0)
    freight_internacional: float = Field(default=0.0)
    impuestos: float = Field(default=0.0)
    valor_total_factura: float = Field(description="Total CPT / Grand Total final con todo incluido.")


class FacturaInternacional(BaseModel):
    encabezado: EncabezadoFactura
    items: List[ItemFactura] = Field(
        description=(
            "Lista de productos UNICOS. REGLA CRITICA: Si un producto aparece 2 veces "
            "(una con descripcion corta y otra con metadatos), SOLO incluirlo UNA VEZ "
            "con la descripcion limpia de la PRIMERA aparicion. "
            "EXCLUIR filas 'CARTON', 'PIWOODENBOX', subtotales, freight, pesos parciales."
        )
    )
    pie: PieFactura


class FacturaProcesada(BaseModel):
    factura: FacturaInternacional
    es_valida: bool = False
    errores: List[str] = Field(default_factory=list)
    advertencias: List[str] = Field(default_factory=list)
    suma_items_calculada: float = 0.0
    fecha_procesamiento: str = Field(default_factory=lambda: datetime.now().isoformat())


# ===========================================================================
# 4. EXTRACCION DE PDF
# ===========================================================================
class PDFExtractor:
    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def extraer_texto(self, pdf_path: str) -> str:
        try:
            reader = PdfReader(pdf_path)
            texto = ""
            for i, page in enumerate(reader.pages, 1):
                txt = page.extract_text() or ""
                texto += f"\n--- PAGE {i} ---\n{txt}\n"
            self.logger.debug(f"PDF '{pdf_path}': {len(reader.pages)} paginas extraidas.")
            return texto.strip()
        except Exception as e:
            self.logger.error(f"Error leyendo PDF {pdf_path}: {e}")
            return ""


# ===========================================================================
# 5. PROCESAMIENTO CON GEMINI (LLM)
# ===========================================================================
class LLMParser:
    def __init__(self, cfg: Config, logger: logging.Logger):
        if not GENAI_DISPONIBLE:
            raise RuntimeError("google-genai no esta instalado.")
        self.client = genai.Client(api_key=cfg.gemini_api_key)
        self.modelo = cfg.modelo_llm
        self.logger = logger
        self.cfg = cfg

    def parsear(self, texto_pdf: str) -> FacturaInternacional:
        prompt = (
            "Analiza el siguiente texto extraido de una Commercial Invoice de Baker Hughes (USA o Mexico).\n"
            "El texto puede estar desordenado porque proviene de un PDF con tablas complejas.\n"
            "Reconstruye la estructura y extrae Encabezado, Items de producto y Pie de pagina.\n\n"
            "=== REGLAS CRITICAS PARA ITEMS ===\n"
            "1. Cada producto aparece DOS VECES en la factura:\n"
            "   - PRIMERA aparicion: codigo + descripcion tecnica limpia + cantidad + precio unitario + extended price.\n"
            "   - SEGUNDA aparicion: la misma descripcion (a veces ligeramente diferente) + metadatos.\n"
            "   - DEBES incluir el producto UNA SOLA VEZ, usando la descripcion de la PRIMERA aparicion.\n"
            "   - De la SEGUNDA aparicion extrae: Delivery, Order, Serial No., commodity codes, ECCN, License Type, Country of Origin.\n\n"
            "2. EXCLUIR filas que NO sean productos con codigo de material:\n"
            "   - Filas 'CARTON' o 'PIWOODENBOX' (son empaques, no productos).\n"
            "   - Filas que solo contengan pesos parciales (Gross Weight, Net Weight, Dimensions) sin codigo de producto.\n"
            "   - Filas de subtotales o freight.\n\n"
            "3. DESCRIPCION LIMPIA:\n"
            "   - Solo el nombre tecnico del producto de la PRIMERA aparicion.\n"
            "   - Ejemplos correctos: 'WINDOW,TEFLON (1673MA)', 'O-RING 234 FKM 95 EL07 BLUE-DOT',\n"
            "     'BEARING DRIVESHAFT UPR 1673MA', 'INSTMT 1677EA ACOUSTIC 8 CHANNEL ELECNS',\n"
            "     'BELLOWS/PRESSURE COMPENSATOR, RUBBER DCBIL', 'COVER TERMINAL MOTOR 562/MSP2-3, M42 SERIES'.\n"
            "   - NO incluir: Delivery, Order, Export/Import commodity code, ECCN, License Type, Country of Origin, Serial No., MEXICAN CUSTOMS IMPORT.\n\n"
            "4. CAMPOS ESPECIFICOS - REGLA DE AUSENCIA:\n"
            "   - Si un campo NO aparece en la factura, DEBES usar EXACTAMENTE el valor por defecto definido en el schema.\n"
            "   - numero_referencia: el valor de 'Delivery: XXXXXX'. Si NO existe: 'NO TIENE DELIVERY'.\n"
            "   - numero_orden: el valor de 'Order: XXXXXX'. Si NO existe: 'NO TIENE ORDEN'.\n"
            "   - numero_serie: el valor de 'Serial No. XXXXXX'. Si NO existe: 'NO TIENE SERIAL'.\n"
            "   - customs_import_mexico: el valor de 'MEXICAN CUSTOMS IMPORT: XXXXXX'. Si NO existe: 'NO APLICA'.\n"
            "   - codigo_exportacion: numero puro del Export country commodity code. Si NO existe: 'NO TIENE CODIGO DE EXPORTACION'.\n"
            "   - codigo_importacion: numero puro del Import country commodity code. Si NO existe: 'NO TIENE CODIGO DE IMPORTACION'.\n"
            "   - pais_origen: pais del Country of Origin. Si NO existe: 'N/A'.\n\n"
            "5. PRECIOS Y PORCENTAJES:\n"
            "   - El precio unitario es el valor que aparece en la columna 'Unit Price'.\n"
            "   - El valor total (Extended Price) es Cantidad x Precio Unitario.\n"
            "   - Si existe un porcentaje de descuento o adicional en el item, incluirlo en el campo 'porcentaje'.\n"
            "   - Si no hay porcentaje, dejarlo en null.\n\n"
            "=== REGLAS PARA EL ENCABEZADO ===\n"
            "- numero_orden: numero de orden general de la factura si existe. Si NO existe: 'NO DISPONIBLE'.\n\n"
            "=== REGLAS PARA EL PIE ===\n"
            "- Net Value: suma pura de todos los Extended Price.\n"
            "- Inland Freight: valor si aparece.\n"
            "- International Freight: valor si aparece.\n"
            "- Total CPT / Grand Total: valor final con todo incluido.\n"
            "- Pesos: extraer LB y KG por separado si aparecen ambos.\n"
            "- Si no aparece freight separado, dejar en 0.\n\n"
            f"TEXTO A PROCESAR:\n{texto_pdf}"
        )

        try:
            response = self.client.models.generate_content(
                model=self.modelo,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=FacturaInternacional,
                    temperature=0.0,
                    system_instruction=(
                        "Eres un auditor contable especializado en facturas de comercio exterior de Baker Hughes. "
                        "Extraes datos con precision quirurgica. Nunca duplicas items. "
                        "Si un campo no existe en la factura, SIEMPRE usas el valor por defecto definido en el schema. "
                        "Nunca dejas strings vacios ni inventas datos."
                    )
                )
            )
            factura = FacturaInternacional.model_validate_json(response.text)
            self.logger.info(f"Factura parseada: {factura.encabezado.numero_factura}")
            return factura
        except ValidationError as e:
            self.logger.error(f"Error validando estructura de factura: {e}")
            raise
        except Exception as e:
            self.logger.error(f"Error en LLM: {e}")
            raise


# ===========================================================================
# 6. SANEAMIENTO DE CAMPOS AUSENTES
# ===========================================================================
class DataSanitizer:
    """
    Fuerza valores explicitos cuando el LLM devuelve strings vacios
    o valores nulos en campos que deben tener un valor por defecto marcado.
    """

    DEFAULTS_ITEM = {
        "numero_referencia": "NO TIENE DELIVERY",
        "numero_orden": "NO TIENE ORDEN",
        "numero_serie": "NO TIENE SERIAL",
        "codigo_exportacion": "NO TIENE CODIGO DE EXPORTACION",
        "codigo_importacion": "NO TIENE CODIGO DE IMPORTACION",
        "pais_origen": "N/A",
        "customs_import_mexico": "NO APLICA",
    }

    DEFAULTS_ENCABEZADO = {
        "numero_orden": "NO DISPONIBLE",
        "pais_destino_final": "Colombia",
        "modo_transporte": "N/A",
        "agente_forwarder": "N/A",
        "fecha_factura": "N/A",
        "moneda": "USD",
    }

    DEFAULTS_PIE = {
        "peso_bruto_total_lb": "N/A",
        "peso_bruto_total_kg": "N/A",
        "peso_neto_total_lb": "N/A",
        "peso_neto_total_kg": "N/A",
    }

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def sanitizar(self, factura: FacturaInternacional) -> FacturaInternacional:
        # Saneamiento de encabezado
        for campo, default in self.DEFAULTS_ENCABEZADO.items():
            valor_actual = getattr(factura.encabezado, campo, None)
            if valor_actual is None or (isinstance(valor_actual, str) and valor_actual.strip() == ""):
                setattr(factura.encabezado, campo, default)
                self.logger.debug(f"Campo '{campo}' vacio en encabezado -> '{default}'")

        # Saneamiento de items
        for item in factura.items:
            for campo, default in self.DEFAULTS_ITEM.items():
                valor_actual = getattr(item, campo, None)
                if valor_actual is None or (isinstance(valor_actual, str) and valor_actual.strip() == ""):
                    setattr(item, campo, default)
                    self.logger.debug(f"Item {item.codigo_producto}: campo '{campo}' vacio -> '{default}'")

        # Saneamiento de pie
        for campo, default in self.DEFAULTS_PIE.items():
            valor_actual = getattr(factura.pie, campo, None)
            if valor_actual is None or (isinstance(valor_actual, str) and valor_actual.strip() == ""):
                setattr(factura.pie, campo, default)
                self.logger.debug(f"Campo '{campo}' vacio en pie -> '{default}'")

        self.logger.info(f"Saneamiento completado para factura {factura.encabezado.numero_factura}")
        return factura


# ===========================================================================
# 7. NORMALIZACION DE DESCRIPCIONES (ETAPA INDEPENDIENTE Y EXTENSIBLE)
# ===========================================================================
class DescriptionNormalizer:
    PATRONES_METADATOS = [
        r"Delivery:\s*\S+",
        r"Order:\s*\S+",
        r"Export country commodity code:\s*\S+",
        r"Import country commodity code:\s*\S+",
        r"ECCN:\s*\S+",
        r"License Type:\s*\S+",
        r"Country of Origin\s*:\s*\S+",
        r"Serial No\.\s*\S+",
        r"MEXICAN CUSTOMS IMPORT:\s*\S+",
        r"\*\*PREPACKAGE.*?\*\*",
    ]

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def normalizar(self, factura: FacturaInternacional) -> FacturaInternacional:
        for item in factura.items:
            item.descripcion_original = item.descripcion_tecnica_limpia
            limpia = item.descripcion_tecnica_limpia

            lineas = limpia.split("\n")
            lineas_limpias = []
            for linea in lineas:
                linea_strip = linea.strip()
                es_metadata = any(re.search(patron, linea_strip, re.IGNORECASE) for patron in self.PATRONES_METADATOS)
                if not es_metadata and linea_strip:
                    lineas_limpias.append(linea_strip)

            limpia = " ".join(lineas_limpias)
            limpia = re.sub(r"\s+", " ", limpia).strip()
            item.descripcion_tecnica_limpia = limpia

        self.logger.debug(f"Normalizacion aplicada a {len(factura.items)} items.")
        return factura

    def aplicar_reglas_custom(self, factura: FacturaInternacional, reglas: Dict[str, Any]) -> FacturaInternacional:
        self.logger.info(f"Aplicando {len(reglas)} reglas custom de descripcion.")
        return factura


# ===========================================================================
# 8. VALIDACION MATEMATICA (DOS NIVELES + ESTADOS)
# ===========================================================================
class InvoiceValidator:
    def __init__(self, cfg: Config, logger: logging.Logger):
        self.cfg = cfg
        self.logger = logger

    def validar(self, factura: FacturaInternacional) -> FacturaProcesada:
        resultado = FacturaProcesada(factura=factura)

        if not factura.items:
            resultado.errores.append("Sin items de producto.")
            self.logger.warning(f"{factura.encabezado.numero_factura}: Sin items.")
            return resultado

        # --- NIVEL 0: Validacion por item ---
        suma_neta = 0.0
        for idx, item in enumerate(factura.items, start=1):
            total_calc = item.cantidad * item.precio_unitario
            if item.porcentaje is not None and item.porcentaje != 0:
                total_calc = total_calc * (1 + item.porcentaje / 100)
            total_calc = round(total_calc, 2)
            total_rep = round(item.valor_total_item, 2)
            diff = abs(total_calc - total_rep)
            suma_neta += total_calc

            if diff <= self.cfg.tolerancia_advertencia:
                estado = EstadoValidacion.CORRECTO
                mensaje = f"Item #{idx}: OK"
            elif diff <= self.cfg.tolerancia_redondeo:
                estado = EstadoValidacion.ADVERTENCIA
                mensaje = f"Item #{idx}: Diferencia por redondeo (${diff:.2f})"
                resultado.advertencias.append(
                    f"Item {item.codigo_producto}: Cant x PU = ${total_calc}, figura ${total_rep}"
                )
            else:
                estado = EstadoValidacion.ERROR
                mensaje = f"Item #{idx}: INCORRECTO. Esperado ${total_calc}, figura ${total_rep}"
                resultado.errores.append(
                    f"Item {item.codigo_producto}: {item.cantidad} x ${item.precio_unitario} = ${total_calc}, "
                    f"pero figura ${total_rep}"
                )

            item.validacion = ResultadoValidacionItem(
                estado=estado,
                total_calculado=total_calc,
                diferencia=diff,
                mensaje=mensaje
            )

            nivel_log = self.logger.info if estado == EstadoValidacion.CORRECTO else self.logger.warning
            nivel_log(f"{factura.encabezado.numero_factura} | {mensaje}")

        # --- NIVEL 1: Suma items vs Net Value ---
        suma_neta = round(suma_neta, 2)
        neto = round(factura.pie.valor_neto, 2)
        resultado.suma_items_calculada = suma_neta

        if neto > 0:
            diff_neto = abs(suma_neta - neto)
            if diff_neto <= self.cfg.tolerancia_redondeo:
                self.logger.info(f"{factura.encabezado.numero_factura}: Net Value OK (${neto})")
            else:
                err = f"NIVEL 1 - Net Value: Suma items (${suma_neta}) != Net Value (${neto})"
                resultado.errores.append(err)
                self.logger.error(f"{factura.encabezado.numero_factura} | {err}")

        # --- NIVEL 2: Net Value + Freight + Impuestos vs Total CPT ---
        freight_total = round(factura.pie.freight_inland + factura.pie.freight_internacional, 2)
        total_esperado = round(neto + freight_total + factura.pie.impuestos, 2)
        total_pie = round(factura.pie.valor_total_factura, 2)
        diff_total = abs(total_esperado - total_pie)

        if diff_total <= self.cfg.tolerancia_redondeo:
            self.logger.info(f"{factura.encabezado.numero_factura}: Total CPT OK (${total_pie})")
        else:
            err = (
                f"NIVEL 2 - Total CPT: Neto(${neto}) + Freight(${freight_total}) + "
                f"Impuestos(${factura.pie.impuestos}) = ${total_esperado}, pero figura ${total_pie}"
            )
            resultado.errores.append(err)
            self.logger.error(f"{factura.encabezado.numero_factura} | {err}")

        resultado.es_valida = len(resultado.errores) == 0
        return resultado


# ===========================================================================
# 9. BASE DE DATOS (SQLite) + PREVENCION DE DUPLICADOS
# ===========================================================================
class DatabaseManager:
    COLUMNAS_ENCABEZADO = [
        "numero_factura", "fecha_factura", "moneda", "emisor_usa",
        "comprador_colombia", "modo_transporte", "agente_forwarder",
        "numero_orden", "pais_destino_final", "peso_bruto_lb",
        "peso_bruto_kg", "peso_neto_lb", "peso_neto_kg", "valor_neto",
        "freight_inland", "freight_internacional", "impuestos",
        "valor_total_pie", "suma_items_calculada", "es_matematicamente_valida",
        "cantidad_errores", "cantidad_advertencias", "fecha_procesamiento"
    ]

    COLUMNAS_ITEMS = [
        "numero_factura", "codigo_producto", "descripcion_original",
        "descripcion_limpia", "cantidad", "unidad_medida", "precio_unitario",
        "valor_total_item", "porcentaje", "estado_validacion",
        "total_calculado", "diferencia", "numero_referencia", "numero_orden",
        "numero_serie", "codigo_exportacion", "codigo_importacion",
        "pais_origen", "customs_import_mexico"
    ]

    def __init__(self, cfg: Config, logger: logging.Logger):
        self.cfg = cfg
        self.logger = logger
        self._init_db()

    def _init_db(self):
        with sqlite3.connect(self.cfg.db_path) as conn:
            c = conn.cursor()
            c.execute("""
                CREATE TABLE IF NOT EXISTS facturas_encabezado (
                    numero_factura TEXT PRIMARY KEY,
                    fecha_factura TEXT,
                    moneda TEXT,
                    emisor_usa TEXT,
                    comprador_colombia TEXT,
                    modo_transporte TEXT,
                    agente_forwarder TEXT,
                    numero_orden TEXT,
                    pais_destino_final TEXT,
                    peso_bruto_lb TEXT,
                    peso_bruto_kg TEXT,
                    peso_neto_lb TEXT,
                    peso_neto_kg TEXT,
                    valor_neto REAL,
                    freight_inland REAL,
                    freight_internacional REAL,
                    impuestos REAL,
                    valor_total_pie REAL,
                    suma_items_calculada REAL,
                    es_matematicamente_valida INTEGER,
                    cantidad_errores INTEGER,
                    cantidad_advertencias INTEGER,
                    fecha_procesamiento TEXT
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS factura_items_cuerpo (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_factura TEXT,
                    codigo_producto TEXT,
                    descripcion_original TEXT,
                    descripcion_limpia TEXT,
                    cantidad REAL,
                    unidad_medida TEXT,
                    precio_unitario REAL,
                    valor_total_item REAL,
                    porcentaje REAL,
                    estado_validacion TEXT,
                    total_calculado REAL,
                    diferencia REAL,
                    numero_referencia TEXT,
                    numero_orden TEXT,
                    numero_serie TEXT,
                    codigo_exportacion TEXT,
                    codigo_importacion TEXT,
                    pais_origen TEXT,
                    customs_import_mexico TEXT,
                    FOREIGN KEY (numero_factura) REFERENCES facturas_encabezado(numero_factura)
                )
            """)
            c.execute("""
                CREATE INDEX IF NOT EXISTS idx_items_factura
                ON factura_items_cuerpo(numero_factura)
            """)
            conn.commit()
        self.logger.info("Base de datos inicializada.")

    def ya_procesada(self, numero_factura: str) -> bool:
        with sqlite3.connect(self.cfg.db_path) as conn:
            c = conn.cursor()
            c.execute("SELECT 1 FROM facturas_encabezado WHERE numero_factura = ?", (numero_factura,))
            return c.fetchone() is not None

    def guardar(self, proc: FacturaProcesada):
        f = proc.factura
        cols_enc = ", ".join(self.COLUMNAS_ENCABEZADO)
        placeholders_enc = ", ".join(["?"] * len(self.COLUMNAS_ENCABEZADO))

        with sqlite3.connect(self.cfg.db_path) as conn:
            c = conn.cursor()
            c.execute(f"""
                INSERT OR REPLACE INTO facturas_encabezado ({cols_enc})
                VALUES ({placeholders_enc})
            """, (
                f.encabezado.numero_factura,
                f.encabezado.fecha_factura,
                f.encabezado.moneda,
                f.encabezado.emisor_usa,
                f.encabezado.comprador_colombia,
                f.encabezado.modo_transporte,
                f.encabezado.agente_forwarder,
                f.encabezado.numero_orden,
                f.encabezado.pais_destino_final,
                f.pie.peso_bruto_total_lb,
                f.pie.peso_bruto_total_kg,
                f.pie.peso_neto_total_lb,
                f.pie.peso_neto_total_kg,
                f.pie.valor_neto,
                f.pie.freight_inland,
                f.pie.freight_internacional,
                f.pie.impuestos,
                f.pie.valor_total_factura,
                proc.suma_items_calculada,
                1 if proc.es_valida else 0,
                len(proc.errores),
                len(proc.advertencias),
                proc.fecha_procesamiento
            ))

            c.execute("DELETE FROM factura_items_cuerpo WHERE numero_factura = ?",
                     (f.encabezado.numero_factura,))

            cols_items = ", ".join(self.COLUMNAS_ITEMS)
            placeholders_items = ", ".join(["?"] * len(self.COLUMNAS_ITEMS))

            for item in f.items:
                c.execute(f"""
                    INSERT INTO factura_items_cuerpo ({cols_items})
                    VALUES ({placeholders_items})
                """, (
                    f.encabezado.numero_factura,
                    item.codigo_producto,
                    item.descripcion_original,
                    item.descripcion_tecnica_limpia,
                    item.cantidad,
                    item.unidad_medida,
                    item.precio_unitario,
                    item.valor_total_item,
                    item.porcentaje,
                    item.validacion.estado.value,
                    item.validacion.total_calculado,
                    item.validacion.diferencia,
                    item.numero_referencia,
                    item.numero_orden,
                    item.numero_serie,
                    item.codigo_exportacion,
                    item.codigo_importacion,
                    item.pais_origen,
                    item.customs_import_mexico
                ))
            conn.commit()
        self.logger.info(f"Guardado en BD: {f.encabezado.numero_factura}")

    def buscar(self, numero_factura: str) -> Dict[str, Any]:
        with sqlite3.connect(self.cfg.db_path) as conn:
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT * FROM facturas_encabezado WHERE numero_factura = ?", (numero_factura,))
            enc = c.fetchone()
            if not enc:
                return {}
            c.execute("SELECT * FROM factura_items_cuerpo WHERE numero_factura = ?", (numero_factura,))
            items = c.fetchall()
            return {"encabezado": dict(enc), "items": [dict(i) for i in items]}


# ===========================================================================
# 10. EXPORTACIONES (CSV / JSON / EXCEL)
# ===========================================================================
class ExportManager:
    def __init__(self, cfg: Config, logger: logging.Logger):
        self.cfg = cfg
        self.logger = logger

    def exportar(self, procesadas: List[FacturaProcesada]):
        if not procesadas:
            return
        self._csv(procesadas)
        self._json(procesadas)
        self._excel(procesadas)

    def _csv(self, procesadas: List[FacturaProcesada]):
        items_rows = []
        enc_rows = []
        for proc in procesadas:
            f = proc.factura
            for item in f.items:
                items_rows.append({
                    "Numero_Factura": f.encabezado.numero_factura,
                    "Codigo_Producto": item.codigo_producto,
                    "Descripcion_Original": item.descripcion_original,
                    "Descripcion_Limpia": item.descripcion_tecnica_limpia,
                    "Cantidad": item.cantidad,
                    "Unidad_Medida": item.unidad_medida,
                    "Precio_Unitario": item.precio_unitario,
                    "Valor_Total_Item": item.valor_total_item,
                    "Porcentaje": item.porcentaje,
                    "Estado_Validacion": item.validacion.estado.value,
                    "Total_Calculado": item.validacion.total_calculado,
                    "Diferencia": item.validacion.diferencia,
                    "Numero_Referencia": item.numero_referencia,
                    "Numero_Orden": item.numero_orden,
                    "Numero_Serie": item.numero_serie,
                    "Cod_Exportacion": item.codigo_exportacion,
                    "Cod_Importacion": item.codigo_importacion,
                    "Pais_Origen": item.pais_origen,
                    "Customs_Import_Mexico": item.customs_import_mexico
                })
            enc_rows.append({
                "Numero_Factura": f.encabezado.numero_factura,
                "Fecha": f.encabezado.fecha_factura,
                "Moneda": f.encabezado.moneda,
                "Emisor_USA": f.encabezado.emisor_usa,
                "Comprador_COL": f.encabezado.comprador_colombia,
                "Modo_Transporte": f.encabezado.modo_transporte,
                "Forwarder": f.encabezado.agente_forwarder,
                "Numero_Orden": f.encabezado.numero_orden,
                "Pais_Destino": f.encabezado.pais_destino_final,
                "Peso_Bruto_LB": f.pie.peso_bruto_total_lb,
                "Peso_Bruto_KG": f.pie.peso_bruto_total_kg,
                "Peso_Neto_LB": f.pie.peso_neto_total_lb,
                "Peso_Neto_KG": f.pie.peso_neto_total_kg,
                "Valor_Neto": f.pie.valor_neto,
                "Freight_Inland": f.pie.freight_inland,
                "Freight_Internacional": f.pie.freight_internacional,
                "Impuestos": f.pie.impuestos,
                "Total_CPT": f.pie.valor_total_factura,
                "Suma_Calculada": proc.suma_items_calculada,
                "Es_Valida": "SI" if proc.es_valida else "NO",
                "Errores": " | ".join(proc.errores) if proc.errores else "",
                "Advertencias": " | ".join(proc.advertencias) if proc.advertencias else ""
            })

        out = Path(self.cfg.directorio_salida)
        pd.DataFrame(items_rows).to_csv(out / "cuerpo_items_facturas.csv", index=False, encoding="utf-8-sig")
        pd.DataFrame(enc_rows).to_csv(out / "encabezados_facturas.csv", index=False, encoding="utf-8-sig")
        self.logger.info("CSV exportados.")

    def _json(self, procesadas: List[FacturaProcesada]):
        datos = [proc.factura.model_dump() for proc in procesadas]
        path = Path(self.cfg.directorio_salida) / "facturas_completas.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        self.logger.info("JSON exportado.")

    def _excel(self, procesadas: List[FacturaProcesada]):
        if not EXCEL_DISPONIBLE:
            self.logger.warning("openpyxl no instalado. Sin exportacion Excel.")
            return

        out = Path(self.cfg.directorio_salida) / "Reporte_Facturas_Comex.xlsx"
        wb = openpyxl.Workbook()

        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        THIN_BORDER = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # --- HOJA 1: ITEMS ---
        ws = wb.active
        ws.title = "Cuerpo de Items"
        headers = [
            "Factura", "Codigo", "Descripcion Limpia", "Cantidad", "Unidad",
            "Precio Unitario", "Valor Total", "Estado Validacion",
            "Total Calculado", "Diferencia",
            "N Ref.", "N Orden", "N Serie",
            "Cod. Export", "Cod. Import", "Pais Origen", "Customs MX"
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for proc in procesadas:
            f = proc.factura
            for item in f.items:
                ws.append([
                    f.encabezado.numero_factura, item.codigo_producto,
                    item.descripcion_tecnica_limpia, item.cantidad,
                    item.unidad_medida, item.precio_unitario,
                    item.valor_total_item, item.validacion.estado.value,
                    item.validacion.total_calculado, item.validacion.diferencia,
                    item.numero_referencia, item.numero_orden, item.numero_serie,
                    item.codigo_exportacion, item.codigo_importacion,
                    item.pais_origen, item.customs_import_mexico
                ])
                ws.cell(row=row_idx, column=6).number_format = '"$"#,##0.00'
                ws.cell(row=row_idx, column=7).number_format = '"$"#,##0.00'
                ws.cell(row=row_idx, column=9).number_format = '"$"#,##0.00'

                estado = item.validacion.estado
                fill = GREEN_FILL if estado == EstadoValidacion.CORRECTO else (
                    YELLOW_FILL if estado == EstadoValidacion.ADVERTENCIA else RED_FILL
                )
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.border = THIN_BORDER
                    if col_idx == 8:
                        c.fill = fill
                        c.font = Font(bold=True)
                        c.alignment = Alignment(horizontal="center")
                row_idx += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 4, 12), 50)

        # --- HOJA 2: RESUMEN ---
        ws2 = wb.create_sheet(title="Resumen Facturas")
        h2 = [
            "N Factura", "Fecha", "Moneda", "Emisor", "Comprador", "Transporte",
            "Forwarder", "N Orden", "Destino", "Neto", "Freight", "Impuestos", "Total CPT",
            "Valida?", "Errores", "Advertencias"
        ]
        ws2.append(h2)
        for col in range(1, len(h2) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for r_idx, proc in enumerate(procesadas, start=2):
            f = proc.factura
            freight = f.pie.freight_inland + f.pie.freight_internacional
            ws2.append([
                f.encabezado.numero_factura, f.encabezado.fecha_factura,
                f.encabezado.moneda, f.encabezado.emisor_usa,
                f.encabezado.comprador_colombia, f.encabezado.modo_transporte,
                f.encabezado.agente_forwarder, f.encabezado.numero_orden,
                f.encabezado.pais_destino_final,
                f.pie.valor_neto, freight, f.pie.impuestos,
                f.pie.valor_total_factura,
                "SI" if proc.es_valida else "NO",
                len(proc.errores), len(proc.advertencias)
            ])
            for col_idx in [10, 11, 12, 13]:
                ws2.cell(row=r_idx, column=col_idx).number_format = '"$"#,##0.00'

            es_v = proc.es_valida
            for col_idx in range(1, len(h2) + 1):
                c = ws2.cell(row=r_idx, column=col_idx)
                c.border = THIN_BORDER
                if col_idx == 14:
                    c.fill = GREEN_FILL if es_v else RED_FILL
                    c.font = Font(bold=True, color="385723" if es_v else "C00000")
                    c.alignment = Alignment(horizontal="center")

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 4, 12), 40)

        # Grafico
        if len(procesadas) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Totales CPT por Factura"
            chart.y_axis.title = "Monto ($)"
            chart.x_axis.title = "Factura"
            chart.width = 16
            chart.height = 10
            data = Reference(ws2, min_col=13, min_row=1, max_row=len(procesadas) + 1)
            cats = Reference(ws2, min_col=1, min_row=2, max_row=len(procesadas) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws2.add_chart(chart, "S2")

        wb.save(out)
        self.logger.info(f"Excel exportado: {out}")


# ===========================================================================
# 11. HISTORIAL JSON
# ===========================================================================
class HistorialManager:
    def __init__(self, cfg: Config, logger: logging.Logger):
        self.cfg = cfg
        self.logger = logger

    def guardar(self, proc: FacturaProcesada):
        path = Path(self.cfg.json_historial)
        historial = []
        if path.exists():
            try:
                with open(path, "r", encoding="utf-8") as f:
                    historial = json.load(f)
            except json.JSONDecodeError:
                historial = []

        f = proc.factura
        registro = {
            "numero_factura": f.encabezado.numero_factura,
            "fecha_factura": f.encabezado.fecha_factura,
            "moneda": f.encabezado.moneda,
            "emisor_usa": f.encabezado.emisor_usa,
            "comprador_colombia": f.encabezado.comprador_colombia,
            "total_items": len(f.items),
            "valor_neto": f.pie.valor_neto,
            "freight_total": f.pie.freight_inland + f.pie.freight_internacional,
            "impuestos": f.pie.impuestos,
            "valor_total_cpt": f.pie.valor_total_factura,
            "suma_calculada": proc.suma_items_calculada,
            "es_valida": proc.es_valida,
            "errores": proc.errores,
            "advertencias": proc.advertencias,
            "fecha_procesamiento": proc.fecha_procesamiento
        }

        historial = [h for h in historial if h.get("numero_factura") != registro["numero_factura"]]
        historial.append(registro)
        historial.sort(key=lambda x: x.get("fecha_factura", ""))

        with open(path, "w", encoding="utf-8") as f_out:
            json.dump(historial, f_out, ensure_ascii=False, indent=4)
        self.logger.debug(f"Historial actualizado: {f.encabezado.numero_factura}")


# ===========================================================================
# 12. REPORTE FINAL
# ===========================================================================
class ReportGenerator:
    def generar(self, procesadas: List[FacturaProcesada], fallidas: List[Tuple[str, str]], logger: logging.Logger):
        total_pdf = len(procesadas) + len(fallidas)
        ok = [p for p in procesadas if p.es_valida and not p.advertencias]
        con_advertencias = [p for p in procesadas if p.advertencias and not p.errores]
        con_errores = [p for p in procesadas if p.errores]

        total_items = sum(len(p.factura.items) for p in procesadas)
        items_inconsistentes = sum(
            1 for p in procesadas for i in p.factura.items
            if i.validacion.estado in (EstadoValidacion.ERROR, EstadoValidacion.ADVERTENCIA)
        )
        facturas_total_mal = [p for p in procesadas if p.errores]

        logger.info("=" * 70)
        logger.info("RESUMEN FINAL DE PROCESAMIENTO")
        logger.info("=" * 70)
        logger.info(f"  Total PDFs encontrados:        {total_pdf}")
        logger.info(f"  Procesadas correctamente:      {len(ok)}")
        logger.info(f"  Con advertencias:              {len(con_advertencias)}")
        logger.info(f"  Con errores:                   {len(con_errores)}")
        logger.info(f"  Fallidas (excepcion):          {len(fallidas)}")
        logger.info(f"  ----------------------------------------")
        logger.info(f"  Total productos extraidos:     {total_items}")
        logger.info(f"  Productos inconsistentes:      {items_inconsistentes}")
        logger.info(f"  Facturas con total CPT mal:    {len(facturas_total_mal)}")
        logger.info("=" * 70)

        if fallidas:
            logger.info("FACTURAS FALLIDAS:")
            for pdf, err in fallidas:
                logger.info(f"   - {os.path.basename(pdf)}: {err}")

        if con_errores:
            logger.info("FACTURAS CON ERRORES:")
            for p in con_errores:
                logger.info(f"   - {p.factura.encabezado.numero_factura}: {' | '.join(p.errores)}")

        if con_advertencias:
            logger.info("FACTURAS CON ADVERTENCIAS:")
            for p in con_advertencias:
                logger.info(f"   - {p.factura.encabezado.numero_factura}: {' | '.join(p.advertencias)}")


# ===========================================================================
# 13. ORQUESTADOR PRINCIPAL
# ===========================================================================
class InvoicePipeline:
    def __init__(self):
        load_dotenv()
        self.cfg = Config()
        self.logger = setup_logging(self.cfg)
        self.extractor = PDFExtractor(self.logger)
        self.parser = LLMParser(self.cfg, self.logger)
        self.sanitizer = DataSanitizer(self.logger)
        self.normalizer = DescriptionNormalizer(self.logger)
        self.validator = InvoiceValidator(self.cfg, self.logger)
        self.db = DatabaseManager(self.cfg, self.logger)
        self.exporter = ExportManager(self.cfg, self.logger)
        self.historial = HistorialManager(self.cfg, self.logger)
        self.reporter = ReportGenerator()

    def run(self):
        archivos = (
            glob.glob(os.path.join(self.cfg.directorio_entrada, "*.pdf")) +
            glob.glob("*.pdf")
        )

        if not archivos:
            self.logger.warning("No se encontraron archivos PDF.")
            return

        self.logger.info(f"Iniciando procesamiento de {len(archivos)} factura(s)...")
        procesadas: List[FacturaProcesada] = []
        fallidas: List[Tuple[str, str]] = []

        for pdf_path in archivos:
            self.logger.info(f"\nProcesando: {pdf_path}")
            try:
                texto = self.extractor.extraer_texto(pdf_path)
                if not texto:
                    fallidas.append((pdf_path, "Sin texto extraido"))
                    continue

                factura = self.parser.parsear(texto)
                if not factura.encabezado.numero_factura:
                    fallidas.append((pdf_path, "Sin numero de factura detectado"))
                    continue

                if self.db.ya_procesada(factura.encabezado.numero_factura):
                    self.logger.warning(
                        f"Factura {factura.encabezado.numero_factura} ya existe en BD. "
                        f"Se actualizara con los nuevos datos."
                    )

                # NUEVO: Saneamiento de campos ausentes
                factura = self.sanitizer.sanitizar(factura)

                factura = self.normalizer.normalizar(factura)
                proc = self.validator.validar(factura)

                self.db.guardar(proc)
                self.historial.guardar(proc)
                procesadas.append(proc)

                if proc.errores:
                    self.logger.warning(f"Factura {proc.factura.encabezado.numero_factura} tiene errores.")

            except Exception as e:
                self.logger.exception(f"Error procesando {pdf_path}")
                fallidas.append((pdf_path, str(e)))
                continue

        if procesadas:
            self.exporter.exportar(procesadas)

        self.reporter.generar(procesadas, fallidas, self.logger)
        self.logger.info("\nPara consultar una factura usa: buscar_factura('NUMERO')")


# ===========================================================================
# 14. FUNCION DE BUSQUEDA (API publica)
# ===========================================================================
def buscar_factura(numero_factura: str):
    cfg = Config()
    logger = setup_logging(cfg)
    db = DatabaseManager(cfg, logger)
    r = db.buscar(numero_factura)

    if not r:
        print(f"\nFactura '{numero_factura}' no encontrada.")
        return

    enc = r["encabezado"]
    print(f"\n{'='*70}")
    print(f"FACTURA: {enc['numero_factura']}")
    print(f"{'='*70}")
    print(f"  Fecha:      {enc['fecha_factura']}")
    print(f"  Moneda:     {enc['moneda']}")
    print(f"  Emisor:     {enc['emisor_usa']}")
    print(f"  Comprador:  {enc['comprador_colombia']}")
    print(f"  Transporte: {enc['modo_transporte']}")
    print(f"  Forwarder:  {enc['agente_forwarder']}")
    print(f"  N Orden:    {enc['numero_orden']}")
    print(f"  Destino:    {enc['pais_destino_final']}")
    print(f"  Peso Bruto: {enc['peso_bruto_lb']} / {enc['peso_bruto_kg']}")
    print(f"  Peso Neto:  {enc['peso_neto_lb']} / {enc['peso_neto_kg']}")
    print(f"  Net Value:  ${enc['valor_neto']}")
    print(f"  Freight:    ${enc['freight_inland'] + enc['freight_internacional']}")
    print(f"  Impuestos:  ${enc['impuestos']}")
    print(f"  Total CPT:  ${enc['valor_total_pie']}")
    print(f"  Valida?:    {'SI' if enc['es_matematicamente_valida'] else 'NO'}")
    print(f"  Errores:    {enc['cantidad_errores']} | Advertencias: {enc['cantidad_advertencias']}")
    print(f"\n  ITEMS ({len(r['items'])}):")
    print(f"  {'-'*60}")
    for it in r["items"]:
        icono = "OK" if it["estado_validacion"] == "CORRECTO" else (
            "WARN" if it["estado_validacion"] == "ADVERTENCIA" else "ERR"
        )
        print(f"  [{icono}] {it['codigo_producto']} | {it['descripcion_limpia'][:50]}...")
        print(f"      {it['cantidad']} {it['unidad_medida']} | PU:${it['precio_unitario']} | Total:${it['valor_total_item']}")
        print(f"      Calc:${it['total_calculado']} | Dif:${it['diferencia']}")
        print(f"      Ref:{it['numero_referencia']} | Ord:{it['numero_orden']} | Ser:{it['numero_serie']}")
        print(f"      Exp:{it['codigo_exportacion']} | Imp:{it['codigo_importacion']} | Orig:{it['pais_origen']} | MX:{it['customs_import_mexico']}")
        print()


# ===========================================================================
# 15. EJECUCION
# ===========================================================================
if __name__ == "__main__":
    pipeline = InvoicePipeline()
    pipeline.run()
